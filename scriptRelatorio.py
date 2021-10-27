import os
import json
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)

### Function Begin    ###
def readXLSX(name, creditCard, count):
    # The source xlsx file is named as source.xlsx
    wb=load_workbook(name)
    # grab the active worksheet
    ws = wb.active
    #Le os dados da planilha
    source = wb["data"]
    for row in source.iter_rows(min_row=2, min_col=4, max_row=source.max_row, max_col=4):
        for cell in row:
            if (cell.value == creditCard):
                count = count + 1
    
    return creditCard, count

def setValuesRow(__sheet, __data):
    #Cria a aba
    __sheet = book.create_sheet("Data", 1) # insert at second position
    #Adiciona um titulo a aba
    __sheet.title = "Data"
    #Cria uma linha com a descrição das colunas
    headers = ['credit card type','Acount']
    # grab the active worksheet
    __sheet = book["Data"]
    #Adiciona as descrições
    __sheet.append(headers)
    #Laço para adicionar os dados a planilha
    for row in __data:
        __sheet.append(row)
    
    #Retorna a planilha criada
    return __sheet

def createChart(__sheet, __data, __title,__col):
    #Cria a lista de referencia do dados
    data = Reference(__sheet, min_col=__col, min_row=2, max_col=__col, max_row=len(__data)+1)
    #Cria a lista de categorias
    categs = Reference(__sheet, min_col=1, min_row=2, max_row=len(__data)+1)

    #Cria o grafico
    chart = BarChart()
    chart.add_data(data=data)
    chart.set_categories(categs)

    chart.legend = None
    chart.y_axis.majorGridlines = None
    chart.varyColors = True
    #Adiciona titulo ao grafico
    chart.title = __title
    #Retorna o grafico
    return chart

def sendEmail(__anexo, __data):
    # criar a integração com o outlook
    outlook = win32.Dispatch('outlook.application')
    
    # criar um email
    email = outlook.CreateItem(0)
    
    __card1 = __data[0][0]
    __card2 = __data[1][0]
    __card3 = __data[2][0]
    
    __cardValor1 = __data[0][1]
    __cardValor2 = __data[1][1]
    __cardValor3 = __data[2][1]
    
    #read file json
    __file = open("data/Properties.json")
    obj = json.load(__file)
    
    # configurar as informações do seu e-mail
    email.To = obj["AddressEmail"]["emailTo"]
    email.CC = obj["AddressEmail"]["emailCC"]
    email.Subject = "Relatorio de Cartões - E-mail Automático"
    email.HTMLBody = f"""
    <p>Prezados(as)</p>
    <p>Segue o resumo dos dados e anexo os dados completos.</p>
    <p>{__card1}: {__cardValor1}</p>
    <p>{__card2}: {__cardValor2}</p>
    <p>{__card3}: {__cardValor3}</p>
    
    <p>Att.</p>
    <p>Equipe ECQ</p>
    """
    email.Attachments.Add("C://Users/lpintanel/Desktop/PythonExcel/"+__anexo)
    email.Send()
    print("Email Enviado")
    return
### Function End    ###

### Function Main   ###
#Cria varialvel dicionario
data = []
count = 0
#Obten o nome dos arquivos da pasta
entries = os.listdir('planilhas/')
#Laço para contabilizar o dados
for entry in entries:
    result = readXLSX('planilhas/'+entry, "mastercard", count)
    count = result[1]
    
data.append(result)

count = 0
for entry in entries:
    result = readXLSX('planilhas/'+entry, "visa", count)
    count = result[1]
    
data.append(result)
    
count = 0
for entry in entries:
    result = readXLSX('planilhas/'+entry, "maestro", count)
    count = result[1]

data.append(result)

book = Workbook()

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=16)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

#Adicionando estilo
book.add_named_style(highlight)
# grab the active worksheet
sheet = book.active
#Ordenando os dados
sorted(data, key=lambda data: data[0])   # sort by date
#Adiciona dados a planilha
sheet = setValuesRow(sheet,data)

#Cria aba na planilha
#ws2 = book.create_sheet("Result",0) # insert at first position
# grab the active worksheet
ws2 = book["Sheet"]
#Adiciona titulo a aba criada
ws2.title = "Result"
#Adiciona estilo a celula
ws2['A1'].style = highlight
#Adiciona valor a celula
ws2['A1'] = "credit card type"
#Adiciona estilo a celula
ws2['B1'].style = highlight
#Adiciona valor a celula
ws2['B1'] = "Acount"
#Adiciona valor a celula
ws2['A2'] = data[0][0]
#Adiciona estilo a celula
ws2['A3'] = data[1][0]
#Adiciona valor a celula
ws2['A4'] = data[2][0]

#Adiciona valor a celula
ws2['B2'] = data[0][1]
#Adiciona valor a celula
ws2['B3'] = data[1][1]
#Adiciona valor a celula
ws2['B4'] = data[2][1]

#Cria o grafico com os dados
chart = createChart(sheet, data, "Totais",2)
#Adiciona grafico a planilha
ws2.add_chart(chart, "B"+str(len(data[0])+3))

#Atribui o nome a variavel
fileName="Relatorio.xlsx"
#Salva os dados em arquivo .xlsx
book.save(fileName)

#sendEmail(fileName, data)
### Function Main End   ###